from openpyxl import load_workbook,Workbook
wbSearch=Workbook()
wbSearch=load_workbook("search.xlsx")#сюда напиши название своего екселя
wsSearch=wbSearch.active

def subscribe(id,activ):
    kolichestvo = wsSearch.max_row
    for i in range(1,kolichestvo):
        value=wsSearch.cell(row=i,column=1).value
        if value==id:
            value_activities=wsSearch.cell(row=value.row,column=5).value
            if value_activities == None:
                wsSearch.cell(row=value.row,column=5).value=activ
            elif value_activities != None:
                row =kolichestvo+1
                wsSearch.cell(row=row,column=5).value=activ
            
