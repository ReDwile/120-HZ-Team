from openpyxl import load_workbook,Workbook
wbSearch=Workbook()
wbSearch=load_workbook("search.xlsx")#сюда напиши название своего екселя
wsSearch=wbSearch.active

def GetUserActs(id):
    kolichestvo = wsSearch.max_row
    for i in range(1,kolichestvo):
        value=wsSearch.cell(row=i,column=1).value
        if value== id:
            activities=[wsSearch.cell(row=value.row,column=5).value]
    else:
        return None