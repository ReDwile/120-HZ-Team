from  openpyxl.utils import get_column_letter
from openpyxl import Workbook,load_workbook

wb = load_workbook('data/test_1.xlsx')
def getacts():
    for sheet in wb:
        vals =[sheet.title]
    return vals


