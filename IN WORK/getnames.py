from openpyxl import load_workbook,Workbook
wbSearch=Workbook()
wbSearch=load_workbook("search.xlsx")#сюда напиши название своего екселя
wsSearch=wbSearch.active

lookfor_1="white_list"
lookfor_2="admin"

kolichestvo = wsSearch.max_row
def getnames(id):
    for i in range(1,kolichestvo): #вместо 21 можно взять любое число
        value=wsSearch.cell(row=i,column=1).value
        if value== id:
            name_status=wsSearch.cell(row=value.row,column=3).value
            lastname_status=wsSearch.cell(row=value.row,column=4).value
            return (name_status+","+lastname_status)
        else:
            return (None)