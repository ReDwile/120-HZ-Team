from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
import random
import telebot
from telebot import types
from bot.classes import *
from bot.config import *

def start_kbd():
    st = types.ReplyKeyboardMarkup()
    row = types.KeyboardButton("/start")
    st.row(row)
    return st


Mes = sm
Man = Person
startkbd = start_kbd()
bot = telebot.TeleBot(f'{TG_TOKEN}')
wb = load_workbook('Data/test.xlsx')
wbSearch=wb
wsSearch=wbSearch.active
lookfor_1="pupil"
lookfor_2="admin"
kolichestvo = wsSearch.max_row


def nonone(a): # Удаление повторений в масссиве
    n = []
    for i in a:
        if i not in n:
            n.append(i)
    return n

def codegen():  #Не допилена
    ran=random.randint(10000, 99999)
    return '$12345'

def db():
    for sheet in wb:
        vals =[sheet.title]
    return vals

def GetUserActs(ID):
    kolichestvo = wsSearch.max_row + 1
    activities=[]
    zero=[]
    for i in range(1, kolichestvo):
        value = wsSearch.cell(row=i,column=1).value
        if value == ID:
            activities.append(wsSearch.cell(row=i, column=5).value)
    nonone(activities)
    for i in range(0, len(activities)):
        if activities[i] == None:
            activities[i] = ""
    if zero == activities:
        return None
    else:
        return nonone(activities)

def getacts():
    kolichestvo = wsSearch.max_row + 1
    activities=[]
    for i in range(1,kolichestvo):
        value = wsSearch.cell(row=i, column=5).value
        if value != "" and value != None:
            activities.append(value)
    return nonone(activities)

def getnames(ID):
    z = ""
    name_status=None
    for i in range(1,kolichestvo+1):
        value=wsSearch.cell(row=i,column=1).value
        if value == ID:
            name_status=wsSearch.cell(row=i,column=3).value
    if name_status!=None:
        return name_status
    return z

def getSnames(ID):
    z = ""
    lastname_status = None
    for i in range(1, kolichestvo + 1):
        value = wsSearch.cell(row=i, column=1).value
        if value == ID:
            lastname_status = wsSearch.cell(row=i, column=4).value
    if lastname_status != None:
        return lastname_status
    return z

def identy(ID):
    status = "noname"
    kolichestvo = wsSearch.max_row + 1
    for i in range(1,kolichestvo):
        value = wsSearch.cell(row=i,column=1).value
        if value == ID:
            return wsSearch.cell(row=i, column=2).value
    return status

def subscribe(message):

    id = message.from_user.id
    useracts = GetUserActs(id)
    activ = message.text
    kolichestvo = wsSearch.max_row + 1
    a = False
    if activ in useracts:
        bot.send_message(id, 'Ошибка! Ты уже зарегистрирован на эту активность',  reply_markup=startkbd)
    else:
        for i in range(1, kolichestvo):
            if wsSearch.cell(row = i, column=1).value == id:
                if wsSearch.cell(row = i, column=5).value == "" or wsSearch.cell(row = kolichestvo, column=5).value == None:
                    a = True
                    RecRow = i
        if a:
            wsSearch.cell(row=RecRow, column=5).value = activ
        else:
            wsSearch.cell(row = kolichestvo, column=1).value = id
            wsSearch.cell(row = kolichestvo, column=5).value = activ
            wsSearch.cell(row = kolichestvo, column=2).value = identy(id)
            wsSearch.cell(row = kolichestvo, column=3).value = getnames(id)
            wsSearch.cell(row = kolichestvo, column=4).value = getSnames(id)
        wb.save("Data/test.xlsx")
        bot.send_message(id, 'Успешно!', reply_markup=startkbd)

def todb():
    maxrow = wsSearch.max_row
    if wsSearch.cell(row=1, column=1).value == "" or wsSearch.cell(row=1, column=1).value == None :
        Row = maxrow
    else:
        Row =maxrow+1
    wsSearch.cell(row=Row, column=1).value = Man.ID
    wsSearch.cell(row=Row, column=2).value = "pupil"
    wsSearch.cell(row=Row, column=3).value = Man.name
    wsSearch.cell(row=Row, column=4).value = Man.lastname
    wb.save("Data/test.xlsx")
    bot.send_message(Man.ID, "Успешно!", reply_markup=startkbd)

def add_act(message):
    name = message.text
    Row = wsSearch.max_row + 1
    wsSearch.cell(row = Row, column = 5).value = name
    wb.save("Data/test.xlsx")
    bot.send_message(message.from_user.id, "Успешно!", reply_markup=startkbd)

def del_act(message):
    name = message.text
    cout = wsSearch.max_row + 1
    for i in range(1, cout):
        if wsSearch.cell(row=i, column=5).value == name:
            wsSearch.cell(row = i, column=5).value = ""
    wb.save("Data/test.xlsx")
    bot.send_message(message.from_user.id, "Успешно!", reply_markup=startkbd)

def send(message):
    group = []
    gr = message.text
    Row = wsSearch.max_row + 1
    if gr == "Всем":
        for i in range(1,Row):
            group.append(wsSearch.cell(row=i, column=1).value)
    else:
        for i in range(1,Row):
            if wsSearch.cell(row = i, column=5).value == gr:
                group.append(wsSearch.cell(row = i, column=1).value)

    bot.send_message(message.from_user.id, "Введите сообщение:")
    bot.register_next_step_handler(message, send2)
    Mes.group = nonone(group)

def send2(message):
    Mes.text = message.text
    for i in range(0, len(Mes.group)):
        bot.send_message(Mes.group[i], Mes.text)
    bot.send_message(message.from_user.id, "Успех!", reply_markup=startkbd)

def checkcode(message):
    inputCode = message.text

    Man.ID = message.from_user.id

    if inputCode == codegen():
        bot.send_message(Man.ID, "Введи свое имя:")
        bot.register_next_step_handler(message, input_name)
    else:
        bot.send_message(Man.ID, text="Ты ввел неверный код!", reply_markup=startkbd)

def desub(message):
    Man.ID = message.from_user.id
    activ = message.text
    n = wsSearch.max_row + 1
    Man.acts = GetUserActs(Man.ID)
    if activ == "Отписаться от всех активностей":
        for i in range(1, n):
            if wsSearch.cell(row=i, column=1).value == Man.ID:
                wsSearch.cell(row=i, column=5).value = ""
                wb.save("Data/test.xlsx")
        bot.send_message(Man.ID, "Успех!", reply_markup=startkbd)
    elif activ in Man.acts:
        for i in range(1, n):
            if wsSearch.cell(i, 1).value == Man.ID:
                if wsSearch.cell(i, 5).value == activ:
                    wsSearch.cell(i, 5).value = ""
                    wb.save("Data/test.xlsx")
        bot.send_message(Man.ID, "Успех!", reply_markup=startkbd)
    else:
        bot.send_message(Man.ID, "Ошибка! Ты не подписан ни на одну активность!", reply_markup=startkbd)


def input_name(message):
    Man.name = message.text
    bot.send_message(Man.ID, "Введи свою фамилию:")
    bot.register_next_step_handler(message, input_surname)

def input_surname(message):
    Man.lastname = message.text
    todb()

