import datetime
from openpyxl import load_workbook



def opredelit():
    wb = load_workbook('Data/test.xlsx')
    wbSearch = wb
    wsSearch = wbSearch.active
    now = datetime.datetime.now()
    value_1 = wsSearch.cell(row=wsSearch.max_row , column=2)
    if now.day > value_1 
        return True
    elif now.day < value_1:
        return True
    else:
        return False
