from openpyxl import load_workbook
from random import randint
import telebot
from telebot import types
from config import *
import datetime
import time
from Telegram.keyboard import PupilTGkeyboards, AdminTGkeyboards, NonameTGKeyboards, backkbd
from modules.getText import getText
import pymysql
from threading import Thread

AdminTexts = getText("admin")
PupilTexts = getText("pupil")
NonameTexts = getText("noname")
OthersTexts = getText("other")

Unknown_command = OthersTexts.gettext("Unknown command")


def start_kbd():
    st = types.ReplyKeyboardMarkup()
    row = types.KeyboardButton("/start")
    st.row(row)
    return st


bot = telebot.TeleBot(f'{TG_TOKEN}')

adminkbd = AdminTGkeyboards.admin_k()
nonamekbd = NonameTGKeyboards.n_k()
pupilkbd = PupilTGkeyboards.p_k()


class MansDataBase(object):
    @staticmethod
    def nonone(a):  # Удаление повторений в масссиве
        n = []
        for i in a:
            if i not in n:
                n.append(i)
        return n

    @staticmethod
    def maxrowSearch(ws):
        if ws.max_row == 1:
            if ws.cell(row=1, column=1).value == "" or ws.cell(row=1, column=1).value == None:
                return 1
            else:
                return 2
        else:
            return int(ws.max_row + 1)

    def __init__(self):
        self.con = pymysql.connect('localhost', 'lalkalol_crochz', 'python', 'lalkalol_crochz')

        self.ID = ""
        self.man = ""
        self.acts = []
        self.VkId = ""


class MansDataBaseTG(MansDataBase):  # Датабаза людей

    def delActs(self, name):  # Функция удаляет активности

        sql = f"UPDATE MainTableV1 SET acts = '' WHERE acts = %s"

        with self.con:
            cur = self.con.cursor()
            cur.execute(sql, name)

    def todb(self, id):  # Функция добавляет пользователя в датабазу
        self.ID = id
        sql = f"INSERT INTO MainTableV1 VALUES(%s, 'pupil','','')"
        with self.con:
            cur = self.con.cursor()
            cur.execute(sql, self.ID)

        bot.send_message(self.ID, "Ты успешно зарегистрирован", reply_markup=pupilkbd)

    def identy(self, ID):  # Определяет статус человека, который написал сообщение
        IdArray = []
        ManArray = []
        status = "noname"
        sql = f"SELECT TgId, Man FROM MainTableV1 WHERE TgId = %s"
        with self.con:
            cur = self.con.cursor()
            cur.execute(sql, ID)
            rows = cur.fetchall()
            for row in rows:
                TgId, Man = row
                IdArray.append(TgId)
                ManArray.append(Man)
        for i in range(0, len(IdArray)):
            status = ManArray[i]
        return status

    def GetUserActs(self, ID):  # Возвращает информацию о всех активностях пользователя
        actsArray = []
        sql = f"SELECT acts FROM MainTableV1 WHERE TgId = %s"
        with self.con:
            cur = self.con.cursor()
            cur.execute(sql, int(ID))
            rows = cur.fetchall()
            for row in rows:
                acts = row[0]
                actsArray.append(acts)

        return self.nonone(actsArray)

    def getVkId(self, ID):
        VkId_status = ""
        sql = f"SELECT VkId FROM MainTableV1 WHERE TgId = %s"
        with self.con:
            cur = self.con.cursor()
            cur.execute(sql, int(ID))
            rows = cur.fetchall()
            for row in rows:
                VkId_status = row[0]

        if VkId_status != 0:
            return VkId_status
        else:
            return "Не указано"

    def subscribe(self, message):  # Подписка на активность
        if message.text == "Назад":
            bot.send_message(message.from_user.id, text=OthersTexts.gettext('back'), reply_markup=pupilkbd)
        else:
            id = message.from_user.id
            useracts = self.GetUserActs(id)
            c = True
            b = True
            if useracts == None:
                b = False
            activ = message.text
            a = False
            if b:
                if activ in useracts:
                    bot.send_message(id, f'Ошибка! Ты уже зарегистрирован на активность «{activ}»',
                                     reply_markup=pupilkbd)
                    c = False
                if not activ in ActsDataBase().getallacts():
                    c = False
                    bot.send_message(id, f'Ошибка! Активности «{activ}» не существует', reply_markup=pupilkbd)
            if c:
                with self.con:
                    cur = self.con.cursor()
                    self.VkId = self.getVkId(self.ID)
                    if self.VkId == "Не указано":
                        self.VkId = ""
                    sql = f"INSERT INTO MainTableV1 VALUES(%s, 'pupil' ,%s , %s)"

                    sqlt = [int(self.ID), activ, self.VkId]

                    cur.execute(sql, sqlt)

                bot.send_message(id, f"Ты подписался на активность «{activ}»", reply_markup=pupilkbd)

    def desub(self, message):  # Отписаться от активности
        if message.text == "Назад":
            bot.send_message(message.from_user.id, text=OthersTexts.gettext('back'), reply_markup=pupilkbd)
        else:
            self.ID = message.from_user.id
            activ = message.text
            self.acts = self.GetUserActs(self.ID)

            self.VkId = self.getVkId(self.ID)
            if self.VkId == "Не указано":
                self.VkId = ""

            if activ == "Отписаться от всех активностей":
                sql = f"DELETE FROM MainTableV1 WHERE TgId = %s"
                sqlt = self.ID
            elif activ in self.acts:
                sql = f"DELETE FROM MainTableV1 WHERE TgId = %s AND acts = %s"
                sqlt = [self.ID, activ]
            else:
                bot.send_message(self.ID, "Ошибка! Ты не подписан на активность «{activ}»", reply_markup=pupilkbd)
                return 0

            with self.con:
                cur = self.con.cursor()
                cur.execute(sql, sqlt)
                if self.identy(self.ID) == "noname":
                    sql = f"INSERT INTO MainTableV1 VALUES(%s, 'pupil','', %s)"
                    sqlt = [self.ID, self.getVkId(self.ID)]
                    cur.execute(sql, sqlt)
                if activ != "Отписаться от всех активностей":
                    bot.send_message(self.ID, f"Ты отписался от активности «{activ}»", reply_markup=pupilkbd)
                else:
                    bot.send_message(self.ID, f"Ты отписался от всех активностей", reply_markup=pupilkbd)

    def changeVK(self, message):
        self.ID = message.from_user.id
        if message.text == "Назад":
            bot.send_message(self.ID, text=OthersTexts.gettext('back'), reply_markup=pupilkbd)
        else:
            if len(message.text) < 8 or len(message.text) > 10:
                bot.send_message(self.ID, "Некорректный ID", reply_markup=pupilkbd)
            else:
                try:
                    self.VkId = int(message.text)
                    with self.con:
                        cur = self.con.cursor()
                        sql = f"UPDATE MainTableV1 SET VkId = %s WHERE TgId = %s"
                        sqlt = [self.VkId, self.ID]
                        cur.execute(sql, sqlt)
                    bot.send_message(self.ID, "Успешная смена ВК", reply_markup=pupilkbd)
                except ValueError:
                    bot.send_message(self.ID, "Некорректный ID", reply_markup=pupilkbd)


class ActsDataBase:  # Датабаза активностей

    @staticmethod
    def nonone(a):  # Удаление повторений в масссиве
        n = []
        for i in a:
            if i not in n:
                n.append(i)
        return n

    def __init__(self):
        self.con = pymysql.connect('localhost', 'lalkalol_crochz', 'python', 'lalkalol_crochz')  # Путь к файлу

        self.Mans = MansDataBaseTG()

    def getallacts(self):  # получаем все активности
        activities = []
        sql = "SELECT * FROM Acts"
        with self.con:
            cur = self.con.cursor()
            cur.execute(sql)
            rows = cur.fetchall()
            for row in rows:
                act = row[0]
                activities.append(act)

        return self.nonone(activities)

    def add_act(self, message):  # Добавляем активность
        if message.text == "Назад":
            bot.send_message(message.from_user.id, text=OthersTexts.gettext('back'), reply_markup=adminkbd)
        else:
            name = message.text
            sql = f"INSERT INTO Acts VALUES(%s)"
            with self.con:
                cur = self.con.cursor()
                cur.execute(sql, name)
            bot.send_message(message.from_user.id, f"Активность «{name}» добавлена", reply_markup=adminkbd)

    def del_act(self, message):  # Удаление активности
        if message.text == "Назад":
            bot.send_message(message.from_user.id, text=OthersTexts.gettext('back'), reply_markup=adminkbd)
        else:
            name = message.text
            sql = f"DELETE FROM Acts WHERE name = %s"

            with self.con:
                cur = self.con.cursor()
                cur.execute(sql, name)

            bot.send_message(message.from_user.id, f"Активность «{name}» удалена", reply_markup=adminkbd)

            self.Mans.delActs(name)  # Удаляем активности из БД людей


class SendMessage:
    @staticmethod
    def nonone(a):  # Удаление повторений в масссиве
        n = []
        for i in a:
            if i not in n:
                n.append(i)
        return n

    def __init__(self):
        self.con = pymysql.connect('localhost', 'lalkalol_crochz', 'python', 'lalkalol_crochz')  # Путь к файлу

        self.text = ""
        self.group = []

    def send(self, message):
        if message.text == "Назад":
            bot.send_message(message.from_user.id, text=OthersTexts.gettext('back'), reply_markup=adminkbd)
        else:
            group = []
            gr = message.text
            if gr == "Всем":
                sql = "SELECT TgId FROM MainTableV1"
            else:
                sql = f"SELECT TgId FROM MainTableV1 WHERE acts = %s"

            with self.con:
                cur = self.con.cursor()
                if gr == "Всем":
                    cur.execute(sql)
                else:
                    cur.execute(sql, gr)
                rows = cur.fetchall()
                for row in rows:
                    group.append(row[0])

            bot.send_message(message.from_user.id, "Введите сообщение:", reply_markup=backkbd())
            bot.register_next_step_handler(message, self.send2, gr)
            self.group = self.nonone(group)

    def send2(self, message, gr):
        if message.text == "Назад":
            bot.send_message(message.from_user.id, text=OthersTexts.gettext('back'), reply_markup=adminkbd)
        else:
            self.text = message.text
            for i in range(0, len(self.group)):
                bot.send_message(self.group[i], text=self.text + '\n\nДля группы: ' + gr)
            bot.send_message(message.from_user.id, f"Сообщение отправлено активности «{gr}»", reply_markup=adminkbd)


class CodeCheck:
    @staticmethod
    def maxrowSearch(ws):
        if ws.max_row == 1:
            if ws.cell(row=1, column=1).value == "" or ws.cell(row=1, column=1).value == None:
                return 1
            else:
                return 2
        else:
            return ws.max_row + 1

    def __init__(self):
        self.path = './Data/code.xlsx'
        self.codewb = load_workbook(self.path)
        self.codeSearch = self.codewb.active
        self.maxrow = self.maxrowSearch(self.codeSearch)
        self.MainDB = MansDataBaseTG()

    def code(self):  # достает код
        value = self.codeSearch.cell(row=self.codeSearch.max_row, column=1).value
        return str(value)

    def checkcode(self, message):
        if message.text == "Назад":
            bot.send_message(message.from_user.id, text=OthersTexts.gettext('back'), reply_markup=nonamekbd)
        else:
            inputCode = message.text
            Man = MansDataBaseTG()

            Man.ID = message.from_user.id
            if inputCode == self.code():
                self.MainDB.todb(Man.ID)
            else:
                bot.send_message(Man.ID, text="Ты ввел неверный код!", reply_markup=nonamekbd)


class CodeGen:

    def __init__(self):
        self.path = './Data/code.xlsx'
        self.codewb = load_workbook(self.path)
        self.codeSearch = self.codewb.active
        self.maxrow = 1
        self.now = datetime.datetime.now()

    def codegen(self):
        code = randint(10000, 99999)
        self.codeSearch.cell(row=self.codeSearch.max_row, column=2).value = self.now.day
        self.codeSearch.cell(row=self.codeSearch.max_row, column=1).value = code
        self.codewb.save(self.path)

    def getcode(self):
        return self.codeSearch.cell(row=self.codeSearch.max_row, column=1).value

    def checkdate(self):
        while True:
            day = self.now.day
            if str(self.codeSearch.cell(row=self.codeSearch.max_row, column=2).value) != str(day):
                self.codegen()
            else:
                pass
            time.sleep(300)
