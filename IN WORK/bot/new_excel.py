from openpyxl import load_workbook
import datetime

def code():
    wb = load_workbook('./test.xlsx')
    wsSearch = wb.active
    value = wsSearch.cell(row=wsSearch.max_row, column=1).value
    return (value)
