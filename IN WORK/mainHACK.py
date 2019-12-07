import telebot
from telebot import types
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
import random

pupil_kbd = types.ReplyKeyboardMarkup()
row1 = types.KeyboardButton("Покажи мои активности")
row2 = types.KeyboardButton("Подписаться на активность")
pupil_kbd.row(row1)
pupil_kbd.row(row2)

nn_kbd = types.ReplyKeyboardMarkup()
row3 = types.KeyboardButton("Регистрация")
nn_kbd.row(row3)

wb = load_workbook('./test.xlsx')
wbSearch=wb
wsSearch=wbSearch.active
lookfor_1="pupil"
lookfor_2="admin"
kolichestvo = wsSearch.max_row

def codegen():
    ran=random.randint(10000, 99999)
    return ('$12345')


def db():
    for sheet in wb:
        vals =[sheet.title]
    return vals

def getacts():
    kolichestvo = wsSearch.max_row
    for i in range(1,kolichestvo):
        value=wsSearch.cell(row=i,column=5).value
        activities = []
        #activities.append(wsSearch.cell(row=value.row,column=5).value)
        set(activities)
        return(activities)

def getnames(ID):
    for i in range(1,kolichestvo):
        value=wsSearch.cell(row=i,column=1).value
        if value== ID:
            value = str(value)
            name_status=wsSearch.cell(row=value.row,column=3).value
            lastname_status=wsSearch.cell(row=value.row,column=4).value
            return (name_status+","+lastname_status)
        else:
            return (None)

def GetUserActs(ID):
    kolichestvo = wsSearch.max_row
    for i in range(1,kolichestvo):
        value=wsSearch.cell(row=i,column=1).value
        zero=[]
        activities=[]
        if value== ID:
            value = str(value)
            activities.append(wsSearch.cell(row=value.row,column=5).value)

        if zero ==activities:
            return None
        else:
            return activities

def identy(ID):
    status = "noname"
    kolichestvo = wsSearch.max_row + 1
    for i in range(1,kolichestvo):
        value=wsSearch.cell(row=i,column=1).value
        if value == ID:
            status = wsSearch.cell(row=i,column=2).value
    return status

def subscribe(id,activ):
    kolichestvo = wsSearch.max_row
    for i in range(1,kolichestvo):
        value=wsSearch.cell(row=i,column=1).value
        if value==id:
            value_activities=wsSearch.cell(row=value.row,column=5).value
            if value_activities == None:
                wsSearch.cell(row=value.row,column=5).value=activ
            elif value_activities != None:
                row =kolichestvo+1
                wsSearch.cell(row=row,column=5).value=activ
    wb.save("test.xlsx")

def todb(ID, name, lastname):
    maxrow = wsSearch.max_row
    Row =maxrow+1
    wsSearch.cell(row=Row, column=1).value = ID
    wsSearch.cell(row=Row, column=3).value = name
    wsSearch.cell(row=Row, column=4).value = lastname
    wb.save("test.xlsx")


TG_TOKEN = "990269257:AAEysundUN5cft2Q0MASpltMA-AiPhKS7nw"

item =[]
activities_kbd = types.ReplyKeyboardMarkup()
acts = getacts()
n = len(acts) - 1
for i in range(0, n, 1):
    item[i] = '/' + types.ReplyKeyboardMarkup(f'{acts[i]}')
    activities_kbd.row(item[i])

class Person:
    name = ""
    lastname = ""
    ID = ""
    man = ""
    acts = []

bot = telebot.TeleBot(f'{TG_TOKEN}')

@bot.message_handler(commands=['start'])
def start(message):
    Man = Person()
    Man.ID = message.from_user.id
    Man.man = identy(Man.ID)

    if Man.man == "noname":
        Man.name = message.from_user.first_name
        Man.lastname = message.from_user.last_name
    else:
        Man.acts = GetUserActs(Man.ID)
        temp = getnames(Man.ID).split(',')
        Man.name = temp[0]
        Man.lastname = temp[1]

    bot.send_message(Man.ID, text=f'{Man.man}')
    if Man.man == "pupil":
        bot.send_message(Man.ID, text="Вот функции, доступные тебе:", reply_markup=pupil_kbd)
    elif Man.man == "noname":
        bot.send_message(Man.ID, text="Вот функции, доступные тебе:", reply_markup=nn_kbd)
    elif Man.man == "admin":
        bot.send_message(Man.ID, text="Вот функции, доступные тебе:", reply_markup=admin_kbd)

@bot.message_handler(content_types=['text'])
def text(message):
    Man = Person()
    Man.ID = message.from_user.id
    Man.man = identy(Man.ID)
    if Man.man == "noname":
        Man.name = message.from_user.first_name
        Man.lastname = message.from_user.last_name
    else:
        Man.acts = GetUserActs(Man.ID)

        temp = getnames(Man.ID).split(',')

        Man.name = temp[0]
        Man.lastname = temp[1]

    if Man.man == "pupil":
        bot.send_message(Man.ID, text="Вот функции, доступные тебе:", reply_markup=pupil_kbd)
    elif Man.man == "admin":
        bot.send_message(Man.ID, text="Вот функции, доступные тебе:", reply_markup=admin_kbd)


    if message.text == "Регистрация":  #Регистрация ноунэйма
        if Man.man != "noname":
            bot.send_message(Man.ID, text="Ты уже зарегистрирован!")
        else:
            bot.send_message(Man.ID, text="Введи код приглашения", reply_markup=pupil_kbd)

    elif message.text[0] == "$":
        if Man.man != "noname":
            bot.send_message(Man.ID, text="Ты уже зарегистрирован!")
        elif message.text == codegen():
            todb(Man.ID,Man.name,Man.lastname)
            bot.send_message(Man.ID, text="Поздравляю, ты зарегистрирован!")
        else:
            bot.send_message(Man.ID, text="Ты ввел неверный код!")

    elif message.text == "Покажи мои активности":
        stru = '\n'.join(Man.acts)
        bot.send_message(Man.ID, text=f"{stru}")

    elif message.text == "Подписаться на активность":
        bot.send_message(Man.ID, text="Держи!\n\n", reply_markup=activities_kbd)

    elif message.text[0] == "/":
        actname = message.text[1:]
        subscribe(Man.ID, actname)

bot.polling(none_stop=True)

