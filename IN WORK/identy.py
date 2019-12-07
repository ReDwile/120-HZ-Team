from openpyxl import load_workbook,Workbook
wbSearch=Workbook()
wbSearch=load_workbook("search.xlsx")#сюда напиши название своего екселя
wsSearch=wbSearch.active

lookfor_1="pupil"
lookfor_2="admin"

def identy(id):
    kolichestvo = wsSearch.max_row
    for i in range(1,kolichestvo):
        value=wsSearch.cell(row=i,column=1).value
        if value== id:
            id_status=wsSearch.cell(row=value.row,column=2)
            if id_status==lookfor_2:
                return("admin")
            elif id_status==lookfor_1:
                return ("pupil")
        else:
            return ("noname")