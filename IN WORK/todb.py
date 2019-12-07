from openpyxl import load_workbook,Workbook
wbSearch=Workbook()
wbSearch=load_workbook('data/.xlsx')#сюда напиши название своего екселя
wsSearch=wbSearch.active

def todb(id,name,lastname):
    maxrow = wsSearch.max_row
    row =maxrow+1
    id_value=wsSearch.cell(row=row, column=1).value=id
    name_value = wsSearch.cell(row=row, column=3).value = name
    lastname_value = wsSearch.cell(row=row, column=4).value = lastname


