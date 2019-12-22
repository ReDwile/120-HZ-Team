from openpyxl import load_workbook
from random import randint
import telebot
from telebot import types
from bot.config import *
import datetime
import time
import bs4
import requests


def start_kbd():
    st = types.ReplyKeyboardMarkup()
    row = types.KeyboardButton("/start")
    st.row(row)
    return st


bot = telebot.TeleBot(f'{TG_TOKEN}')

class MansDataBase(object):
    @staticmethod
    def nonone(a):  # Удаление повторений в масссиве
        n = []
        for i in a:
            if i not in n:
                n.append(i)
        return n

    @staticmethod
    def maxrowSearch(ws):
        if ws.max_row == 1:
            if ws.cell(row=1, column=1).value == "" or ws.cell(row=1, column=1).value == None:
                return 1
            else:
                return 2
        else:
            return int(ws.max_row + 1)

    def __init__(self):
        self.path = '/Users/lalkalol/Desktop/bot/Data/test.xlsx'  # Путь к файлу
        self.wb = load_workbook(self.path)  # Создаем книгу
        self.wsSearch = self.wb.active  # Рабочий лист
        self.maxrow = self.maxrowSearch(self.wsSearch)  # Узнаем максимальный ряд

        self.name = ""
        self.lastname = ""
        self.ID = ""
        self.man = ""
        self.acts = []
        self.VkId = ""

class MansDataBaseTG(MansDataBase):  # Датабаза людей

    def delActs(self, name):  # Функция удаляет активности
        for i in range(1, self.maxrow):
            if self.wsSearch.cell(row=i, column=5).value == name:
                self.wsSearch.cell(row=i, column=5).value = ""
        self.wb.save(self.path)

    def todb(self):  # Функция добавляет пользователя в датабазу
        self.wsSearch.cell(row=self.maxrow, column=1).value = self.ID
        self.wsSearch.cell(row=self.maxrow, column=2).value = "pupil"
        self.wsSearch.cell(row=self.maxrow, column=3).value = self.name
        self.wsSearch.cell(row=self.maxrow, column=4).value = self.lastname
        self.wb.save(self.path)
        bot.send_message(self.ID, "Успешно!", reply_markup=startkbd)

    def identy(self, ID):  # Определяет статус человека, который написал сообщение
        status = "noname"
        for i in range(1, self.maxrow):
            value = self.wsSearch.cell(row=i, column=1).value
            if value == ID:
                return self.wsSearch.cell(row=i, column=2).value
        return status

    def GetUserActs(self, ID):  # Возвращает информацию о всех активностях пользователя
        activities = []
        zero = []
        for i in range(1, self.maxrow):
            value = self.wsSearch.cell(row=i, column=1).value
            if value == ID:
                activities.append(self.wsSearch.cell(row=i, column=5).value)
        self.nonone(activities)
        for i in range(0, len(activities)):
            if activities[i] == None:
                activities[i] = ""
        if zero == activities:
            return None
        else:
            return self.nonone(activities)

    def getnames(self, ID):  # Возвращает имя пользователя
        z = ""
        name_status = None
        for i in range(1, self.maxrow):
            value = self.wsSearch.cell(row=i, column=1).value
            if value == ID:
                name_status = self.wsSearch.cell(row=i, column=3).value
        if name_status != None:
            return name_status
        return z

    def getSnames(self, ID):  # Возвращает фамилию пользователя
        z = ""
        lastname_status = None
        for i in range(1, self.maxrow):
            value = self.wsSearch.cell(row=i, column=1).value
            if value == ID:
                lastname_status = self.wsSearch.cell(row=i, column=4).value
        if lastname_status != None:
            return lastname_status
        return z

    def getVkId(self,ID):
        z = ""
        VkId_status = None
        for i in range(1, self.maxrow):
            value = self.wsSearch.cell(row=i, column=1).value
            if value == ID:
                VkId_status = self.wsSearch.cell(row=i, column=46).value
        if VkId_status != None:
            return VkId_status
        return z

    def subscribe(self, message):  # Подписка на активность
        id = message.from_user.id
        useracts = self.GetUserActs(id)
        activ = message.text
        a = False
        if activ in useracts:
            bot.send_message(id, 'Ошибка! Ты уже зарегистрирован на эту активность', reply_markup=startkbd)
        else:
            for i in range(1, self.maxrow):
                if self.wsSearch.cell(row=i, column=1).value == id:
                    if self.wsSearch.cell(row=i, column=5).value == "" or self.wsSearch.cell(row=i, column=5).value == None:
                        a = True
                        RecRow = i
            if a:
                self.wsSearch.cell(row=RecRow, column=5).value = activ
            else:
                self.wsSearch.cell(row=self.maxrow, column=1).value = id
                self.wsSearch.cell(row=self.maxrow, column=5).value = activ
                self.wsSearch.cell(row=self.maxrow, column=2).value = self.identy(id)
                self.wsSearch.cell(row=self.maxrow, column=3).value = self.getnames(id)
                self.wsSearch.cell(row=self.maxrow, column=4).value = self.getSnames(id)
                self.wsSearch.cell(row=self.maxrow, column=6).value = self.getVkId(id)
            self.wb.save(self.path)
            bot.send_message(id, 'Успешно!', reply_markup=startkbd)

    def desub(self, message):  # Отписаться от активности
        Man = MansDataBaseTG()

        Man.ID = message.from_user.id
        activ = message.text
        Man.acts = self.GetUserActs(Man.ID)
        if activ == "Отписаться от всех активностей":
            for i in range(1, self.maxrow):
                if self.wsSearch.cell(row=i, column=1).value == Man.ID:
                    self.wsSearch.cell(row=i, column=5).value = ""
                    self.wb.save(self.path)
            bot.send_message(Man.ID, "Успех!", reply_markup=startkbd)
        elif activ in Man.acts:
            for i in range(1, self.maxrow):
                if self.wsSearch.cell(i, 1).value == Man.ID:
                    if self.wsSearch.cell(i, 5).value == activ:
                        self.wsSearch.cell(i, 5).value = ""
                        self.wb.save(self.path)
            bot.send_message(Man.ID, "Успех!", reply_markup=startkbd)
        else:
            bot.send_message(Man.ID, "Ошибка! Ты не подписан ни на одну активность!", reply_markup=startkbd)

    def input_name(self, message):  # Ввод имени при регистрации
        self.ID = message.from_user.id
        self.name = message.text
        bot.send_message(message.from_user.id, "Введи свою фамилию:")
        bot.register_next_step_handler(message, self.input_surname)

    def input_surname(self, message):  # Ввод фамилию при регистрации
        self.lastname = message.text
        self.todb()

    def addVk(self, message):  # Добавляем вк
        self.ID = message.from_user.id
        self.VkId = message.text
        for i in range(1, self.maxrow):
            if self.ID == self.wsSearch.cell(row=i, column=1).value:
                self.wsSearch.cell(row=i, column=6).value = self.VkId
        self.wb.save(self.path)
        bot.send_message(self.ID, "Успешно!", reply_markup=startkbd)

class MansDataBaseVK(MansDataBase):
    def todbVK(self, message):
        self.VkId = message.user_id
        self.name = self.get_VKuser_name(self.VkId)
        self.lastname = self.get_VKuser_lastname(self.VkId)

        self.wsSearch.cell(row=self.maxrow, column=6).value = self.VkId
        self.wsSearch.cell(row=self.maxrow, column=2).value = "pupil"
        self.wsSearch.cell(row=self.maxrow, column=3).value = self.name
        self.wsSearch.cell(row=self.maxrow, column=4).value = self.lastname


    def get_VKuser_lastname(self, user_id):
        request = requests.get("https://vk.com/" + str(user_id))
        bs = bs4.BeautifulSoup(request.text, "html.parser")

        user_name = self._clean_all_tag_from_str(bs.findAll("title")[0])

        return user_name.split()[1]

    def identyVK(self, ID):
        status = "noname"
        for i in range(1, self.maxrow):
            value = self.wsSearch.cell(row=i, column=6).value
            if value == ID:
                return self.wsSearch.cell(row=i, column=2).value
        return status

    def get_VKuser_name(self, user_id):
        request = requests.get("https://vk.com/" + str(user_id))
        bs = bs4.BeautifulSoup(request.text, "html.parser")

        user_name = self._clean_all_tag_from_str(bs.findAll("title")[0])
        return user_name.split()[0]

    def GetUserActsVK(self, ID):
        activities = []
        zero = []
        for i in range(6, self.maxrow):
            value = self.wsSearch.cell(row=i, column=6).value
            if value == ID:
                activities.append(self.wsSearch.cell(row=i, column=5).value)
        self.nonone(activities)
        for i in range(0, len(activities)):
            if activities[i] == None:
                activities[i] = ""
        if zero == activities:
            return None
        else:
            return self.nonone(activities)

    @staticmethod
    def _clean_all_tag_from_str(string_line):
        """
        Очистка строки stringLine от тэгов и их содержимых
        :param string_line: Очищаемая строка
        :return: очищенная строка
        """
        result = ""
        not_skip = True
        for i in list(string_line):
            if not_skip:
                if i == "<":
                    not_skip = False
                else:
                    result += i
            else:
                if i == ">":
                    not_skip = True

        return result


class ActsDataBase:  # Датабаза активностей
    @staticmethod
    def nonone(a):  # Удаление повторений в масссиве
        n = []
        for i in a:
            if i not in n:
                n.append(i)
        return n

    @staticmethod
    def maxrowSearch(ws):
        if ws.max_row == 1:
            if ws.cell(row=1, column=1).value == "" or ws.cell(row=1, column=1).value == None:
                return 1
            else:
                return 2
        else:
            return ws.max_row + 1

    def __init__(self):
        self.path = '/Users/lalkalol/Desktop/bot/Data/acts.xlsx'  # Путь к файлу
        self.wb = load_workbook(self.path)  # Создаем книгу
        self.wsSearch = self.wb.active  # Рабочий лист

        self.maxrow = self.maxrowSearch(self.wsSearch)  # Узнаем максимальный ряд

        self.Mans = MansDataBaseTG()

    def getallacts(self):  # получаем все активности
        activities = []
        for i in range(1, self.maxrow):
            value = self.wsSearch.cell(row=i, column=1).value
            if value != "" and value != None:
                activities.append(value)
        return self.nonone(activities)

    def add_act(self, message):  # Добавляем активность
        name = message.text
        self.wsSearch.cell(row=self.maxrow, column=1).value = name
        self.wb.save(self.path)
        bot.send_message(message.from_user.id, "Успешно!", reply_markup=startkbd)

    def del_act(self, message):  # Удаление активности
        name = message.text
        for i in range(1, self.maxrow):
            if self.wsSearch.cell(row=i, column=1).value == name:
                self.wsSearch.cell(row=i, column=1).value = ""
        self.wb.save(self.path)
        bot.send_message(message.from_user.id, "Успешно!", reply_markup=startkbd)

        self.Mans.delActs(name)  # Удаляем активности из БД людей


class SendMessage:
    @staticmethod
    def nonone(a):  # Удаление повторений в масссиве
        n = []
        for i in a:
            if i not in n:
                n.append(i)
        return n

    @staticmethod
    def maxrowSearch(ws):
        if ws.max_row == 1:
            if ws.cell(row=1, column=1).value == "" or ws.cell(row=1, column=1).value == None:
                return 1
            else:
                return 2
        else:
            return ws.max_row + 1

    def __init__(self):
        self.path = '/Users/lalkalol/Desktop/bot/Data/test.xlsx'  # Путь к файлу
        self.wb = load_workbook(self.path)  # Создаем книгу
        self.wsSearch = self.wb.active  # Рабочий лист
        self.maxrow = self.maxrowSearch(self.wsSearch)  # Узнаем максимальный ряд

        self.text = ""
        self.group = []

    def send(self, message):
        group = []
        gr = message.text
        if gr == "Всем":
            for i in range(1, self.maxrow):
                group.append(self.wsSearch.cell(row=i, column=1).value)
        else:
            for i in range(1, self.maxrow):
                if self.wsSearch.cell(row=i, column=5).value == gr:
                    group.append(self.wsSearch.cell(row=i, column=1).value)

        bot.send_message(message.from_user.id, "Введите сообщение:")
        bot.register_next_step_handler(message, self.send2)
        self.group = self.nonone(group)

    def send2(self, message):
        self.text = message.text
        for i in range(0, len(self.group)):
            bot.send_message(self.group[i], self.text)
        bot.send_message(message.from_user.id, "Успех!", reply_markup=startkbd)


class CodeCheck:
    @staticmethod
    def maxrowSearch(ws):
        if ws.max_row == 1:
            if ws.cell(row=1, column=1).value == "" or ws.cell(row=1, column=1).value == None:
                return 1
            else:
                return 2
        else:
            return ws.max_row + 1

    def __init__(self):
        self.path = '/Users/lalkalol/Desktop/bot/Data/code.xlsx'
        self.codewb = load_workbook(self.path)
        self.codeSearch = self.codewb.active
        self.maxrow = self.maxrowSearch(self.codeSearch)
        self.MainDB = MansDataBaseTG()

    def code(self):  # достает код
        value = self.codeSearch.cell(row=self.codeSearch.max_row, column=1).value
        return str(value)

    def checkcode(self, message):
        inputCode = message.text
        Man = MansDataBaseTG()

        Man.ID = message.from_user.id
        if True:
            if inputCode == self.code():
                bot.send_message(Man.ID, "Введи свое имя:")
                bot.register_next_step_handler(message, self.MainDB.input_name)
            else:
                bot.send_message(Man.ID, text="Ты ввел неверный код!", reply_markup=startkbd)


class CodeGen:
    @staticmethod
    def maxrowSearch(ws):
        if ws.max_row == 1:
            if ws.cell(row=1, column=1).value == "" or ws.cell(row=1, column=1).value == None:
                return 1
            else:
                return 2
        else:
            return ws.max_row + 1

    def __init__(self):
        self.path = '/Users/lalkalol/Desktop/bot/Data/code.xlsx'
        self.codewb = load_workbook(self.path)
        self.codeSearch = self.codewb.active
        self.maxrow = self.maxrowSearch(self.codeSearch)
        self.now = datetime.datetime.now()

    def checkdate(self):
        while True:
            day = self.now.day
            if self.codeSearch.cell(row=self.codeSearch.max_row, column=2) != str(day):
                self.codegen()
            time.sleep(1800)

    def codegen(self):
        code = randint(10000, 99999)
        self.codeSearch.cell(row=self.codeSearch.max_row, column=2).value = self.now.day
        self.codeSearch.cell(row=self.codeSearch.max_row, column=1).value = code
        self.codewb.save(self.path)


startkbd = start_kbd()

